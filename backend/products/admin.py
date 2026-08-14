from decimal import Decimal, InvalidOperation
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from django.contrib import admin, messages
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.db import IntegrityError, transaction
from django.shortcuts import redirect, render
from django.urls import path
from django.utils.html import format_html
from django.utils.text import slugify

from openpyxl import load_workbook

from categories.models import (
    Category,
    Department,
    SubCategory,
)

from .forms import (
    ProductBulkUploadForm,
    ProductImageBulkUploadForm,
    ProductVariantBulkUploadForm,
)
from .models import (
    Brand,
    Product,
    ProductImage,
    ProductVariant,
)


# =========================================================
# Common Helpers
# =========================================================

TRUE_VALUES = {
    "true",
    "1",
    "yes",
    "y",
    "on",
}

FALSE_VALUES = {
    "false",
    "0",
    "no",
    "n",
    "off",
}


def clean_text(value):
    if value is None:
        return ""

    return str(value).strip()


def clean_boolean(
    value,
    default=False,
):
    if value is None or value == "":
        return default

    if isinstance(value, bool):
        return value

    normalized = (
        str(value)
        .strip()
        .lower()
    )

    if normalized in TRUE_VALUES:
        return True

    if normalized in FALSE_VALUES:
        return False

    return default


def clean_decimal(
    value,
    allow_empty=False,
):
    if value is None or value == "":
        return None

    try:
        decimal_value = Decimal(
            str(value).strip()
        )

        if decimal_value < 0:
            return None

        return decimal_value

    except (
        InvalidOperation,
        ValueError,
        TypeError,
    ):
        return None


def clean_integer(
    value,
    default=0,
    minimum=0,
):
    if value is None or value == "":
        return default

    try:
        number = int(
            value
        )

        return max(
            number,
            minimum,
        )

    except (
        TypeError,
        ValueError,
    ):
        return default


def validation_error_message(
    error,
):
    if isinstance(
        error,
        ValidationError,
    ):
        if hasattr(
            error,
            "message_dict",
        ):
            messages_list = []

            for field, errors in (
                error.message_dict.items()
            ):
                messages_list.append(
                    f"{field}: "
                    + ", ".join(
                        str(item)
                        for item in errors
                    )
                )

            return "; ".join(
                messages_list
            )

        return "; ".join(
            str(item)
            for item in error.messages
        )

    return str(
        error
    )


# =========================================================
# Product Image Inline
# =========================================================

class ProductImageInline(
    admin.TabularInline
):
    model = ProductImage

    extra = 1

    fields = (
        "image",
        "alt_text",
        "order",
    )

    ordering = (
        "order",
        "id",
    )


# =========================================================
# Product Variant Inline
# =========================================================

class ProductVariantInline(
    admin.TabularInline
):
    model = ProductVariant

    extra = 1

    fields = (
        "color",
        "color_code",
        "size",
        "stock",
        "sku",
        "is_active",
    )

    ordering = (
        "color",
        "size",
    )


# =========================================================
# Brand Admin
# =========================================================

@admin.register(Brand)
class BrandAdmin(
    admin.ModelAdmin
):
    list_display = (
        "name",
        "slug",
        "is_active",
    )

    list_editable = (
        "is_active",
    )

    list_filter = (
        "is_active",
    )

    search_fields = (
        "name",
        "slug",
    )

    prepopulated_fields = {
        "slug": (
            "name",
        ),
    }

    ordering = (
        "name",
    )

    save_on_top = True


# =========================================================
# Product Admin
# =========================================================

@admin.register(Product)
class ProductAdmin(
    admin.ModelAdmin
):
    change_list_template = (
        "admin/products/product/change_list.html"
    )

    # -----------------------------------------------------
    # Product List
    # -----------------------------------------------------

    list_display = (
        "name",
        "sku",
        "brand",
        "department",
        "category",
        "subcategory",
        "price",
        "old_price",
        "admin_total_stock",
        "admin_stock_status",
        "is_active",
        "is_featured",
        "is_best_seller",
        "is_trending",
        "is_new_arrival",
        "is_clearance_sale",
        "is_offer",
    )

    list_editable = (
        "is_active",
        "is_featured",
        "is_best_seller",
        "is_trending",
        "is_new_arrival",
        "is_clearance_sale",
        "is_offer",
    )

    list_filter = (
        "department",
        "category",
        "subcategory",
        "brand",
        "is_active",
        "is_featured",
        "is_best_seller",
        "is_trending",
        "is_new_arrival",
        "is_clearance_sale",
        "is_offer",
    )

    search_fields = (
        "name",
        "slug",
        "sku",
        "brand__name",
        "department__name",
        "category__name",
        "subcategory__name",
        "description",
    )

    ordering = (
        "-created_at",
    )

    list_select_related = (
        "brand",
        "department",
        "category",
        "subcategory",
    )

    list_per_page = 50

    save_on_top = True

    date_hierarchy = (
        "created_at"
    )

    prepopulated_fields = {
        "slug": (
            "name",
        ),
    }

    readonly_fields = (
        "rating",
        "total_reviews",
        "admin_total_stock_detail",
        "created_at",
        "updated_at",
    )

    inlines = (
        ProductImageInline,
        ProductVariantInline,
    )

    # -----------------------------------------------------
    # Product Edit Page
    # -----------------------------------------------------

    fieldsets = (
        (
            "Product Information",
            {
                "fields": (
                    "name",
                    "slug",
                    "sku",
                    "description",
                    "brand",
                ),
            },
        ),
        (
            "Category",
            {
                "fields": (
                    "department",
                    "category",
                    "subcategory",
                ),
            },
        ),
        (
            "Pricing",
            {
                "fields": (
                    "price",
                    "old_price",
                ),
            },
        ),
        (
            "Images",
            {
                "fields": (
                    "main_image",
                ),
            },
        ),
        (
            "Inventory Summary",
            {
                "fields": (
                    "admin_total_stock_detail",
                ),
            },
        ),
        (
            "Store Collections",
            {
                "fields": (
                    "is_active",
                    "is_featured",
                    "is_best_seller",
                    "is_trending",
                    "is_new_arrival",
                    "is_clearance_sale",
                    "is_offer",
                ),
            },
        ),
        (
            "Reviews",
            {
                "fields": (
                    "rating",
                    "total_reviews",
                ),
            },
        ),
        (
            "SEO",
            {
                "classes": (
                    "collapse",
                ),
                "fields": (
                    "meta_title",
                    "meta_description",
                ),
            },
        ),
        (
            "System Information",
            {
                "classes": (
                    "collapse",
                ),
                "fields": (
                    "created_at",
                    "updated_at",
                ),
            },
        ),
    )

    # =====================================================
    # Stock Display
    # =====================================================

    @admin.display(
        description="Stock",
    )
    def admin_total_stock(
        self,
        obj,
    ):
        return obj.total_stock

    @admin.display(
        description="Availability",
    )
    def admin_stock_status(
        self,
        obj,
    ):
        if obj.in_stock:
            return format_html(
                '<span style="color:#198754;font-weight:600;">'
                "In stock"
                "</span>"
            )

        return format_html(
            '<span style="color:#dc3545;font-weight:600;">'
            "Out of stock"
            "</span>"
        )

    @admin.display(
        description="Total Active Variant Stock",
    )
    def admin_total_stock_detail(
        self,
        obj,
    ):
        if not obj.pk:
            return 0

        return obj.total_stock

    # =====================================================
    # Query Optimization
    # =====================================================

    def get_queryset(
        self,
        request,
    ):
        queryset = (
            super()
            .get_queryset(
                request
            )
        )

        return (
            queryset
            .select_related(
                "brand",
                "department",
                "category",
                "subcategory",
            )
            .prefetch_related(
                "variants",
            )
        )

    # =====================================================
    # Custom Admin URLs
    # =====================================================

    def get_urls(self):
        default_urls = (
            super().get_urls()
        )

        custom_urls = [
            path(
                "bulk-upload/",
                self.admin_site.admin_view(
                    self.bulk_upload_view
                ),
                name=(
                    "products_product_"
                    "bulk_upload"
                ),
            ),
            path(
                "image-bulk-upload/",
                self.admin_site.admin_view(
                    self.image_bulk_upload_view
                ),
                name=(
                    "products_product_"
                    "image_bulk_upload"
                ),
            ),
        ]

        return (
            custom_urls
            + default_urls
        )

    # =====================================================
    # Bulk Product Upload Page
    # =====================================================

    def bulk_upload_view(
        self,
        request,
    ):
        if request.method == "POST":
            form = ProductBulkUploadForm(
                request.POST,
                request.FILES,
            )

            if form.is_valid():
                excel_file = (
                    form.cleaned_data[
                        "excel_file"
                    ]
                )

                try:
                    result = (
                        self.import_products(
                            excel_file
                        )
                    )

                    level = (
                        messages.WARNING
                        if result[
                            "skipped"
                        ]
                        else messages.SUCCESS
                    )

                    self.message_user(
                        request,
                        (
                            "Product upload completed. "
                            f"Created: "
                            f"{result['created']}, "
                            f"Updated: "
                            f"{result['updated']}, "
                            f"Skipped: "
                            f"{result['skipped']}."
                        ),
                        level=level,
                    )

                    return redirect(
                        "admin:"
                        "products_product_"
                        "changelist"
                    )

                except Exception as error:
                    self.message_user(
                        request,
                        (
                            "Product upload "
                            f"failed: {error}"
                        ),
                        level=messages.ERROR,
                    )

        else:
            form = (
                ProductBulkUploadForm()
            )

        context = {
            **self.admin_site.each_context(
                request
            ),
            "title":
                "Bulk Product Upload",
            "form":
                form,
            "opts":
                self.model._meta,
        }

        return render(
            request,
            (
                "admin/products/product/"
                "bulk_upload.html"
            ),
            context,
        )

    # =====================================================
    # Bulk Product Image Upload Page
    # =====================================================

    def image_bulk_upload_view(
        self,
        request,
    ):
        if request.method == "POST":
            form = (
                ProductImageBulkUploadForm(
                    request.POST,
                    request.FILES,
                )
            )

            if form.is_valid():
                zip_file = (
                    form.cleaned_data[
                        "zip_file"
                    ]
                )

                try:
                    result = (
                        self.import_product_images(
                            zip_file
                        )
                    )

                    level = (
                        messages.WARNING
                        if result[
                            "skipped"
                        ]
                        else messages.SUCCESS
                    )

                    self.message_user(
                        request,
                        (
                            "Image upload completed. "
                            f"Main images: "
                            f"{result['main_images']}, "
                            f"Gallery images: "
                            f"{result['gallery_images']}, "
                            f"Skipped: "
                            f"{result['skipped']}."
                        ),
                        level=level,
                    )

                    return redirect(
                        "admin:"
                        "products_product_"
                        "changelist"
                    )

                except Exception as error:
                    self.message_user(
                        request,
                        (
                            "Image upload "
                            f"failed: {error}"
                        ),
                        level=messages.ERROR,
                    )

        else:
            form = (
                ProductImageBulkUploadForm()
            )

        context = {
            **self.admin_site.each_context(
                request
            ),
            "title":
                "Bulk Product Image Upload",
            "form":
                form,
            "opts":
                self.model._meta,
        }

        return render(
            request,
            (
                "admin/products/product/"
                "image_bulk_upload.html"
            ),
            context,
        )

    # =====================================================
    # Product Excel Import
    # =====================================================

    def import_products(
        self,
        excel_file,
    ):
        workbook = load_workbook(
            excel_file,
            read_only=True,
            data_only=True,
        )

        try:
            sheet = workbook.active

            rows = sheet.iter_rows(
                values_only=True
            )

            try:
                raw_headers = next(
                    rows
                )
            except StopIteration as error:
                raise ValueError(
                    "Excel file is empty."
                ) from error

            headers = [
                clean_text(
                    header
                ).lower()
                for header
                in raw_headers
            ]

            required_headers = {
                "sku",
                "name",
                "brand",
                "department",
                "category",
                "price",
            }

            missing_headers = (
                required_headers
                .difference(
                    headers
                )
            )

            if missing_headers:
                raise ValueError(
                    (
                        "Missing required "
                        "columns: "
                    )
                    + ", ".join(
                        sorted(
                            missing_headers
                        )
                    )
                )

            created_count = 0
            updated_count = 0
            skipped_count = 0

            for (
                row_number,
                row,
            ) in enumerate(
                rows,
                start=2,
            ):
                row_data = dict(
                    zip(
                        headers,
                        row,
                    )
                )

                if not any(
                    value
                    not in (
                        None,
                        "",
                    )
                    for value
                    in row_data.values()
                ):
                    continue

                try:
                    with transaction.atomic():
                        sku = clean_text(
                            row_data.get(
                                "sku"
                            )
                        )

                        name = clean_text(
                            row_data.get(
                                "name"
                            )
                        )

                        brand_name = clean_text(
                            row_data.get(
                                "brand"
                            )
                        )

                        department_name = (
                            clean_text(
                                row_data.get(
                                    "department"
                                )
                            )
                        )

                        category_name = (
                            clean_text(
                                row_data.get(
                                    "category"
                                )
                            )
                        )

                        subcategory_name = (
                            clean_text(
                                row_data.get(
                                    "subcategory"
                                )
                            )
                        )

                        if not all(
                            (
                                sku,
                                name,
                                brand_name,
                                department_name,
                                category_name,
                            )
                        ):
                            raise ValueError(
                                (
                                    "Required values "
                                    "are missing."
                                )
                            )

                        price = clean_decimal(
                            row_data.get(
                                "price"
                            )
                        )

                        if price is None:
                            raise ValueError(
                                "Invalid price."
                            )

                        old_price = (
                            clean_decimal(
                                row_data.get(
                                    "old_price"
                                ),
                                allow_empty=True,
                            )
                        )

                        if (
                            old_price
                            is not None
                            and old_price
                            < price
                        ):
                            raise ValueError(
                                (
                                    "Old price cannot "
                                    "be lower than "
                                    "selling price."
                                )
                            )

                        # ---------------------------------
                        # Brand
                        # ---------------------------------

                        brand = (
                            Brand.objects
                            .filter(
                                name__iexact=(
                                    brand_name
                                )
                            )
                            .first()
                        )

                        if brand is None:
                            brand = Brand(
                                name=brand_name,
                                slug=(
                                    self
                                    .unique_brand_slug(
                                        brand_name
                                    )
                                ),
                                is_active=True,
                            )

                            brand.full_clean()
                            brand.save()

                        # ---------------------------------
                        # Department
                        # ---------------------------------

                        department = (
                            Department.objects
                            .filter(
                                name__iexact=(
                                    department_name
                                )
                            )
                            .first()
                        )

                        if department is None:
                            department = (
                                Department(
                                    name=(
                                        department_name
                                    ),
                                    slug=(
                                        self
                                        .unique_department_slug(
                                            department_name
                                        )
                                    ),
                                    is_active=True,
                                    show_in_navbar=True,
                                )
                            )

                            department.full_clean()
                            department.save()

                        # ---------------------------------
                        # Category
                        # ---------------------------------

                        category = (
                            Category.objects
                            .filter(
                                department=(
                                    department
                                ),
                                name__iexact=(
                                    category_name
                                ),
                            )
                            .first()
                        )

                        if category is None:
                            category = Category(
                                department=(
                                    department
                                ),
                                name=(
                                    category_name
                                ),
                                slug=(
                                    self
                                    .unique_category_slug(
                                        department_name,
                                        category_name,
                                    )
                                ),
                                is_active=True,
                                show_in_navbar=True,
                            )

                            category.full_clean()
                            category.save()

                        # ---------------------------------
                        # Subcategory
                        # ---------------------------------

                        subcategory = None

                        if subcategory_name:
                            subcategory = (
                                SubCategory.objects
                                .filter(
                                    category=(
                                        category
                                    ),
                                    name__iexact=(
                                        subcategory_name
                                    ),
                                )
                                .first()
                            )

                            if subcategory is None:
                                subcategory = (
                                    SubCategory(
                                        category=(
                                            category
                                        ),
                                        name=(
                                            subcategory_name
                                        ),
                                        slug=(
                                            self
                                            .unique_subcategory_slug(
                                                department_name,
                                                category_name,
                                                subcategory_name,
                                            )
                                        ),
                                        is_active=True,
                                        show_in_navbar=True,
                                    )
                                )

                                subcategory.full_clean()
                                subcategory.save()

                        # ---------------------------------
                        # Product
                        # ---------------------------------

                        product = (
                            Product.objects
                            .filter(
                                sku=sku
                            )
                            .first()
                        )

                        created = (
                            product is None
                        )

                        if created:
                            product = Product(
                                sku=sku
                            )

                        product_slug = (
                            clean_text(
                                row_data.get(
                                    "slug"
                                )
                            )
                        )

                        if (
                            product_slug
                        ):
                            product.slug = (
                                product_slug
                            )

                        elif (
                            created
                            or not product.slug
                        ):
                            product.slug = (
                                self
                                .unique_product_slug(
                                    name,
                                    sku,
                                )
                            )

                        product.name = name
                        product.brand = brand
                        product.department = (
                            department
                        )
                        product.category = (
                            category
                        )
                        product.subcategory = (
                            subcategory
                        )

                        product.description = (
                            clean_text(
                                row_data.get(
                                    "description"
                                )
                            )
                        )

                        product.price = price
                        product.old_price = (
                            old_price
                        )

                        product.is_active = (
                            clean_boolean(
                                row_data.get(
                                    "is_active"
                                ),
                                default=True,
                            )
                        )

                        product.is_featured = (
                            clean_boolean(
                                row_data.get(
                                    "is_featured"
                                )
                            )
                        )

                        product.is_best_seller = (
                            clean_boolean(
                                row_data.get(
                                    "is_best_seller"
                                )
                            )
                        )

                        product.is_trending = (
                            clean_boolean(
                                row_data.get(
                                    "is_trending"
                                )
                            )
                        )

                        product.is_new_arrival = (
                            clean_boolean(
                                row_data.get(
                                    "is_new_arrival"
                                )
                            )
                        )

                        product.is_clearance_sale = (
                            clean_boolean(
                                row_data.get(
                                    "is_clearance_sale"
                                )
                            )
                        )

                        product.is_offer = (
                            clean_boolean(
                                row_data.get(
                                    "is_offer"
                                )
                            )
                        )

                        product.meta_title = (
                            clean_text(
                                row_data.get(
                                    "meta_title"
                                )
                            )
                        )

                        product.meta_description = (
                            clean_text(
                                row_data.get(
                                    "meta_description"
                                )
                            )
                        )

                        product.full_clean()
                        product.save()

                        if created:
                            created_count += 1
                        else:
                            updated_count += 1

                except Exception as error:
                    skipped_count += 1

                    print(
                        (
                            "Product row "
                            f"{row_number} "
                            "skipped: "
                            f"{validation_error_message(error)}"
                        )
                    )

            return {
                "created":
                    created_count,
                "updated":
                    updated_count,
                "skipped":
                    skipped_count,
            }

        finally:
            workbook.close()

    # =====================================================
    # Product ZIP Image Import
    # =====================================================

    @transaction.atomic
    def import_product_images(
        self,
        zip_file,
    ):
        allowed_extensions = {
            ".jpg",
            ".jpeg",
            ".png",
            ".webp",
        }

        main_image_count = 0
        gallery_image_count = 0
        skipped_count = 0

        try:
            archive = ZipFile(
                zip_file
            )

        except BadZipFile as error:
            raise ValueError(
                (
                    "Uploaded file is "
                    "not a valid ZIP file."
                )
            ) from error

        products_by_sku = {
            product.sku
            .strip()
            .lower():
                product
            for product
            in Product.objects.all()
            if product.sku
        }

        try:
            for zip_info in (
                archive.infolist()
            ):
                if zip_info.is_dir():
                    continue

                original_name = (
                    Path(
                        zip_info.filename
                    ).name
                )

                if not original_name:
                    continue

                if original_name.startswith(
                    "."
                ):
                    continue

                file_path = Path(
                    original_name
                )

                extension = (
                    file_path
                    .suffix
                    .lower()
                )

                if (
                    extension
                    not in allowed_extensions
                ):
                    skipped_count += 1
                    continue

                file_stem = (
                    file_path.stem
                )

                product = None
                gallery_order = None

                exact_sku = (
                    file_stem.lower()
                )

                if (
                    exact_sku
                    in products_by_sku
                ):
                    product = (
                        products_by_sku[
                            exact_sku
                        ]
                    )

                elif "_" in file_stem:
                    (
                        possible_sku,
                        possible_order,
                    ) = file_stem.rsplit(
                        "_",
                        1,
                    )

                    if (
                        possible_order
                        .isdigit()
                        and (
                            possible_sku
                            .lower()
                            in products_by_sku
                        )
                    ):
                        product = (
                            products_by_sku[
                                possible_sku
                                .lower()
                            ]
                        )

                        gallery_order = (
                            int(
                                possible_order
                            )
                        )

                if product is None:
                    print(
                        (
                            "Image skipped: "
                            f"{original_name}. "
                            "Matching product "
                            "SKU not found."
                        )
                    )

                    skipped_count += 1
                    continue

                try:
                    image_bytes = (
                        archive.read(
                            zip_info
                        )
                    )

                    if not image_bytes:
                        skipped_count += 1
                        continue

                    image_content = (
                        ContentFile(
                            image_bytes
                        )
                    )

                    # -----------------------------
                    # Main image
                    # -----------------------------

                    if gallery_order is None:
                        if product.main_image:
                            product.main_image.delete(
                                save=False
                            )

                        product.main_image.save(
                            original_name,
                            image_content,
                            save=True,
                        )

                        main_image_count += 1

                    # -----------------------------
                    # Gallery image
                    # -----------------------------

                    else:
                        existing_image = (
                            ProductImage.objects
                            .filter(
                                product=product,
                                order=gallery_order,
                            )
                            .first()
                        )

                        if existing_image:
                            if (
                                existing_image
                                .image
                            ):
                                (
                                    existing_image
                                    .image
                                    .delete(
                                        save=False
                                    )
                                )

                            existing_image.alt_text = (
                                product.name
                            )

                            existing_image.order = (
                                gallery_order
                            )

                            (
                                existing_image
                                .image
                                .save(
                                    original_name,
                                    image_content,
                                    save=False,
                                )
                            )

                            existing_image.full_clean()
                            existing_image.save()

                        else:
                            gallery_image = (
                                ProductImage(
                                    product=(
                                        product
                                    ),
                                    alt_text=(
                                        product.name
                                    ),
                                    order=(
                                        gallery_order
                                    ),
                                )
                            )

                            (
                                gallery_image
                                .image
                                .save(
                                    original_name,
                                    image_content,
                                    save=False,
                                )
                            )

                            gallery_image.full_clean()
                            gallery_image.save()

                        gallery_image_count += 1

                except Exception as error:
                    print(
                        (
                            "Image "
                            f"{original_name} "
                            "skipped: "
                            f"{error}"
                        )
                    )

                    skipped_count += 1

        finally:
            archive.close()

        return {
            "main_images":
                main_image_count,
            "gallery_images":
                gallery_image_count,
            "skipped":
                skipped_count,
        }

    # =====================================================
    # Slug Helpers
    # =====================================================

    @staticmethod
    def unique_product_slug(
        name,
        sku,
    ):
        base_slug = (
            slugify(
                name
            )
            or "product"
        )

        candidate = base_slug
        counter = 1

        while (
            Product.objects
            .filter(
                slug=candidate
            )
            .exclude(
                sku=sku
            )
            .exists()
        ):
            candidate = (
                f"{base_slug}-"
                f"{counter}"
            )

            counter += 1

        return candidate

    @staticmethod
    def unique_brand_slug(
        name,
    ):
        base_slug = (
            slugify(
                name
            )
            or "brand"
        )

        candidate = base_slug
        counter = 1

        while Brand.objects.filter(
            slug=candidate
        ).exists():
            candidate = (
                f"{base_slug}-"
                f"{counter}"
            )

            counter += 1

        return candidate

    @staticmethod
    def unique_department_slug(
        name,
    ):
        base_slug = (
            slugify(
                name
            )
            or "department"
        )

        candidate = base_slug
        counter = 1

        while (
            Department.objects
            .filter(
                slug=candidate
            )
            .exists()
        ):
            candidate = (
                f"{base_slug}-"
                f"{counter}"
            )

            counter += 1

        return candidate

    @staticmethod
    def unique_category_slug(
        department_name,
        category_name,
    ):
        base_slug = (
            slugify(
                (
                    f"{department_name}-"
                    f"{category_name}"
                )
            )
            or "category"
        )

        candidate = base_slug
        counter = 1

        while (
            Category.objects
            .filter(
                slug=candidate
            )
            .exists()
        ):
            candidate = (
                f"{base_slug}-"
                f"{counter}"
            )

            counter += 1

        return candidate

    @staticmethod
    def unique_subcategory_slug(
        department_name,
        category_name,
        subcategory_name,
    ):
        base_slug = (
            slugify(
                (
                    f"{department_name}-"
                    f"{category_name}-"
                    f"{subcategory_name}"
                )
            )
            or "subcategory"
        )

        candidate = base_slug
        counter = 1

        while (
            SubCategory.objects
            .filter(
                slug=candidate
            )
            .exists()
        ):
            candidate = (
                f"{base_slug}-"
                f"{counter}"
            )

            counter += 1

        return candidate


# =========================================================
# Product Image Admin
# =========================================================

@admin.register(ProductImage)
class ProductImageAdmin(
    admin.ModelAdmin
):
    list_display = (
        "product",
        "image_preview",
        "alt_text",
        "order",
    )

    list_filter = (
        "product",
    )

    list_editable = (
        "order",
    )

    search_fields = (
        "product__name",
        "product__sku",
        "alt_text",
    )

    ordering = (
        "product",
        "order",
        "id",
    )

    list_select_related = (
        "product",
    )

    @admin.display(
        description="Preview",
    )
    def image_preview(
        self,
        obj,
    ):
        if not obj.image:
            return "—"

        try:
            return format_html(
                (
                    '<img src="{}" '
                    'style="width:50px;'
                    'height:50px;'
                    'object-fit:cover;'
                    'border-radius:6px;" />'
                ),
                obj.image.url,
            )

        except ValueError:
            return "—"


# =========================================================
# Product Variant Admin
# =========================================================

@admin.register(ProductVariant)
class ProductVariantAdmin(
    admin.ModelAdmin
):
    change_list_template = (
        "admin/products/productvariant/"
        "change_list.html"
    )

    list_display = (
        "product",
        "color",
        "color_code",
        "size",
        "stock",
        "sku",
        "admin_stock_status",
        "is_active",
    )

    list_editable = (
        "stock",
        "is_active",
    )

    list_filter = (
        "product",
        "color",
        "size",
        "is_active",
    )

    search_fields = (
        "product__name",
        "product__sku",
        "sku",
        "color",
        "size",
    )

    ordering = (
        "product",
        "color",
        "size",
    )

    list_select_related = (
        "product",
    )

    list_per_page = 100

    save_on_top = True

    @admin.display(
        description="Availability",
    )
    def admin_stock_status(
        self,
        obj,
    ):
        if obj.in_stock:
            return format_html(
                '<span style="color:#198754;font-weight:600;">'
                "In stock"
                "</span>"
            )

        return format_html(
            '<span style="color:#dc3545;font-weight:600;">'
            "Out of stock"
            "</span>"
        )

    # =====================================================
    # Custom URLs
    # =====================================================

    def get_urls(self):
        default_urls = (
            super().get_urls()
        )

        custom_urls = [
            path(
                "bulk-upload/",
                self.admin_site.admin_view(
                    self.bulk_upload_view
                ),
                name=(
                    "products_productvariant_"
                    "bulk_upload"
                ),
            ),
        ]

        return (
            custom_urls
            + default_urls
        )

    # =====================================================
    # Variant Upload Page
    # =====================================================

    def bulk_upload_view(
        self,
        request,
    ):
        if request.method == "POST":
            form = (
                ProductVariantBulkUploadForm(
                    request.POST,
                    request.FILES,
                )
            )

            if form.is_valid():
                excel_file = (
                    form.cleaned_data[
                        "excel_file"
                    ]
                )

                try:
                    result = (
                        self.import_variants(
                            excel_file
                        )
                    )

                    level = (
                        messages.WARNING
                        if result[
                            "skipped"
                        ]
                        else messages.SUCCESS
                    )

                    self.message_user(
                        request,
                        (
                            "Variant upload completed. "
                            f"Created: "
                            f"{result['created']}, "
                            f"Updated: "
                            f"{result['updated']}, "
                            f"Skipped: "
                            f"{result['skipped']}."
                        ),
                        level=level,
                    )

                    return redirect(
                        "admin:"
                        "products_productvariant_"
                        "changelist"
                    )

                except Exception as error:
                    self.message_user(
                        request,
                        (
                            "Variant upload failed: "
                            f"{error}"
                        ),
                        level=messages.ERROR,
                    )

        else:
            form = (
                ProductVariantBulkUploadForm()
            )

        context = {
            **self.admin_site.each_context(
                request
            ),
            "title":
                "Bulk Product Variant Upload",
            "form":
                form,
            "opts":
                self.model._meta,
        }

        return render(
            request,
            (
                "admin/products/productvariant/"
                "bulk_upload.html"
            ),
            context,
        )

    # =====================================================
    # Variant Excel Import
    # =====================================================

    def import_variants(
        self,
        excel_file,
    ):
        workbook = load_workbook(
            excel_file,
            read_only=True,
            data_only=True,
        )

        try:
            sheet = workbook.active

            rows = sheet.iter_rows(
                values_only=True
            )

            try:
                raw_headers = next(
                    rows
                )
            except StopIteration as error:
                raise ValueError(
                    "Excel file is empty."
                ) from error

            headers = [
                clean_text(
                    header
                ).lower()
                for header
                in raw_headers
            ]

            required_headers = {
                "product_sku",
                "variant_sku",
                "color",
                "size",
                "stock",
            }

            missing_headers = (
                required_headers
                .difference(
                    headers
                )
            )

            if missing_headers:
                raise ValueError(
                    (
                        "Missing required "
                        "columns: "
                    )
                    + ", ".join(
                        sorted(
                            missing_headers
                        )
                    )
                )

            created_count = 0
            updated_count = 0
            skipped_count = 0

            for (
                row_number,
                row,
            ) in enumerate(
                rows,
                start=2,
            ):
                row_data = dict(
                    zip(
                        headers,
                        row,
                    )
                )

                if not any(
                    value
                    not in (
                        None,
                        "",
                    )
                    for value
                    in row_data.values()
                ):
                    continue

                try:
                    with transaction.atomic():
                        product_sku = (
                            clean_text(
                                row_data.get(
                                    "product_sku"
                                )
                            )
                        )

                        variant_sku = (
                            clean_text(
                                row_data.get(
                                    "variant_sku"
                                )
                            )
                        )

                        color = clean_text(
                            row_data.get(
                                "color"
                            )
                        )

                        color_code = (
                            clean_text(
                                row_data.get(
                                    "color_code"
                                )
                            )
                        )

                        size = clean_text(
                            row_data.get(
                                "size"
                            )
                        )

                        if not all(
                            (
                                product_sku,
                                variant_sku,
                                color,
                                size,
                            )
                        ):
                            raise ValueError(
                                (
                                    "Required values "
                                    "are missing."
                                )
                            )

                        product = (
                            Product.objects
                            .filter(
                                sku__iexact=(
                                    product_sku
                                )
                            )
                            .first()
                        )

                        if product is None:
                            raise ValueError(
                                (
                                    "Product SKU "
                                    f"{product_sku} "
                                    "not found."
                                )
                            )

                        stock_raw = (
                            row_data.get(
                                "stock"
                            )
                        )

                        try:
                            stock = int(
                                stock_raw or 0
                            )

                        except (
                            TypeError,
                            ValueError,
                        ) as error:
                            raise ValueError(
                                (
                                    "Invalid stock "
                                    "value."
                                )
                            ) from error

                        if stock < 0:
                            raise ValueError(
                                (
                                    "Stock cannot "
                                    "be negative."
                                )
                            )

                        is_active = (
                            clean_boolean(
                                row_data.get(
                                    "is_active"
                                ),
                                default=True,
                            )
                        )

                        # ---------------------------------
                        # Prefer existing SKU
                        # ---------------------------------

                        variant = (
                            ProductVariant.objects
                            .filter(
                                sku=variant_sku
                            )
                            .first()
                        )

                        # ---------------------------------
                        # Otherwise find same combination
                        # ---------------------------------

                        if variant is None:
                            variant = (
                                ProductVariant.objects
                                .filter(
                                    product=product,
                                    color__iexact=(
                                        color
                                    ),
                                    size__iexact=(
                                        size
                                    ),
                                )
                                .first()
                            )

                        created = (
                            variant is None
                        )

                        if created:
                            variant = (
                                ProductVariant()
                            )

                        variant.product = product
                        variant.sku = variant_sku
                        variant.color = color
                        variant.color_code = (
                            color_code
                        )
                        variant.size = size
                        variant.stock = stock
                        variant.is_active = (
                            is_active
                        )

                        # This validates the new
                        # product/color/size uniqueness.
                        variant.full_clean()

                        variant.save()

                        if created:
                            created_count += 1
                        else:
                            updated_count += 1

                except (
                    ValidationError,
                    IntegrityError,
                    ValueError,
                ) as error:
                    skipped_count += 1

                    print(
                        (
                            "Variant row "
                            f"{row_number} "
                            "skipped: "
                            f"{validation_error_message(error)}"
                        )
                    )

                except Exception as error:
                    skipped_count += 1

                    print(
                        (
                            "Variant row "
                            f"{row_number} "
                            "skipped: "
                            f"{error}"
                        )
                    )

            return {
                "created":
                    created_count,
                "updated":
                    updated_count,
                "skipped":
                    skipped_count,
            }

        finally:
            workbook.close()