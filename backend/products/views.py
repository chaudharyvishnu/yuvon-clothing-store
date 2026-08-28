from decimal import Decimal, InvalidOperation
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from django.core.files.base import ContentFile
from django.db import IntegrityError, transaction
from django.db.models import Avg, Count, Prefetch, Q
from django.utils.text import slugify

from openpyxl import load_workbook

from rest_framework import generics, permissions, status
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from categories.models import (
    Category,
    Department,
    SubCategory,
)
from reviews.models import Review

from .models import (
    Brand,
    Product,
    ProductImage,
    ProductVariant,
)

from .serializers import (
    BrandSerializer,
    ProductAdminWriteSerializer,
    ProductDetailSerializer,
    ProductImageAdminSerializer,
    ProductListSerializer,
    ProductVariantAdminSerializer,
)


# =========================================================
# Helpers
# =========================================================

TRUE_VALUES = {
    "1",
    "true",
    "yes",
    "on",
    "y",
}

FALSE_VALUES = {
    "0",
    "false",
    "no",
    "off",
    "n",
}


def clean_text(value):
    if value is None:
        return ""

    return str(value).strip()


def clean_boolean(
    value,
    default=False,
):
    if value is None or value == "":
        return default

    normalized = (
        str(value)
        .strip()
        .lower()
    )

    if normalized in TRUE_VALUES:
        return True

    if normalized in FALSE_VALUES:
        return False

    return default


def clean_decimal(
    value,
    allow_empty=False,
):
    if value in {
        None,
        "",
    }:
        if allow_empty:
            return None

        return None

    try:
        return Decimal(
            str(value)
        )

    except (
        InvalidOperation,
        TypeError,
        ValueError,
    ):
        return None


def query_bool(value):
    if value is None:
        return False

    return (
        str(value)
        .strip()
        .lower()
        in TRUE_VALUES
    )


def query_decimal(value):
    return clean_decimal(
        value,
        allow_empty=True,
    )


def unique_slug_for_model(
    model_class,
    base_value,
    exclude_id=None,
):
    base_slug = (
        slugify(
            base_value
        )
        or "item"
    )

    candidate = base_slug
    counter = 2

    queryset = (
        model_class.objects
        .all()
    )

    if exclude_id:
        queryset = (
            queryset.exclude(
                id=exclude_id
            )
        )

    while queryset.filter(
        slug=candidate
    ).exists():
        candidate = (
            f"{base_slug}-{counter}"
        )

        counter += 1

    return candidate


def error_message(error):
    if hasattr(
        error,
        "message_dict",
    ):
        return str(
            error.message_dict
        )

    if hasattr(
        error,
        "messages",
    ):
        return " ".join(
            str(message)
            for message
            in error.messages
        )

    return str(error)


# =========================================================
# Permissions
# =========================================================

class IsAdminUserForProducts(
    permissions.BasePermission
):
    """
    Allow access only to authenticated admin/staff users.
    """

    message = (
        "Admin access is required "
        "to manage products."
    )

    def has_permission(
        self,
        request,
        view,
    ):
        user = request.user

        if (
            not user
            or not user.is_authenticated
        ):
            return False

        return bool(
            getattr(
                user,
                "is_staff",
                False,
            )
            or getattr(
                user,
                "is_superuser",
                False,
            )
            or getattr(
                user,
                "is_admin",
                False,
            )
        )


# =========================================================
# Shared Public Product Queryset
# =========================================================

def base_product_queryset():
    approved_reviews = (
        Review.objects
        .filter(
            status="approved",
        )
        .select_related(
            "user",
            "order_item",
        )
        .order_by(
            "-created_at"
        )
    )

    active_variants = (
        ProductVariant.objects
        .filter(
            is_active=True,
        )
        .order_by(
            "color",
            "size",
        )
    )

    ordered_images = (
        ProductImage.objects
        .all()
        .order_by(
            "order",
            "id",
        )
    )

    return (
        Product.objects
        .filter(
            is_active=True,
        )
        .select_related(
            "brand",
            "department",
            "category",
            "subcategory",
        )
        .prefetch_related(
            Prefetch(
                "variants",
                queryset=active_variants,
            ),
            Prefetch(
                "images",
                queryset=ordered_images,
            ),
            Prefetch(
                "reviews",
                queryset=approved_reviews,
            ),
        )
        .annotate(
            approved_average_rating=Avg(
                "reviews__rating",
                filter=Q(
                    reviews__status="approved"
                ),
            ),
            approved_reviews_count=Count(
                "reviews",
                filter=Q(
                    reviews__status="approved"
                ),
                distinct=True,
            ),
        )
    )


# =========================================================
# Shared Admin Product Queryset
# =========================================================

def admin_product_queryset():
    """
    Admin queryset includes active and inactive products,
    variants, images and category relations.
    """

    return (
        Product.objects
        .all()
        .select_related(
            "brand",
            "department",
            "category",
            "subcategory",
        )
        .prefetch_related(
            "variants",
            "images",
        )
        .order_by(
            "-created_at"
        )
    )


# =========================================================
# Brand List
# =========================================================

class BrandListView(
    generics.ListAPIView
):
    serializer_class = (
        BrandSerializer
    )

    permission_classes = (
        permissions.AllowAny,
    )

    def get_queryset(self):
        return (
            Brand.objects
            .filter(
                is_active=True,
            )
            .order_by(
                "name"
            )
        )


# =========================================================
# Product List
# =========================================================

class ProductListView(
    generics.ListAPIView
):
    serializer_class = (
        ProductListSerializer
    )

    permission_classes = (
        permissions.AllowAny,
    )

    def get_queryset(self):
        queryset = (
            base_product_queryset()
        )

        params = (
            self.request.query_params
        )

        # -------------------------------------------------
        # Search
        # -------------------------------------------------

        search = (
            params
            .get(
                "search",
                "",
            )
            .strip()
        )

        if search:
            queryset = (
                queryset.filter(
                    Q(
                        name__icontains=search
                    )
                    | Q(
                        description__icontains=search
                    )
                    | Q(
                        sku__icontains=search
                    )
                    | Q(
                        brand__name__icontains=search
                    )
                    | Q(
                        department__name__icontains=search
                    )
                    | Q(
                        category__name__icontains=search
                    )
                    | Q(
                        subcategory__name__icontains=search
                    )
                )
            )

        # -------------------------------------------------
        # Navigation filters
        # -------------------------------------------------

        department = (
            params.get(
                "department"
            )
        )

        category = (
            params.get(
                "category"
            )
        )

        subcategory = (
            params.get(
                "subcategory"
            )
        )

        brand = (
            params.get(
                "brand"
            )
        )

        if department:
            queryset = (
                queryset.filter(
                    department__slug=department
                )
            )

        if category:
            queryset = (
                queryset.filter(
                    category__slug=category
                )
            )

        if subcategory:
            queryset = (
                queryset.filter(
                    subcategory__slug=subcategory
                )
            )

        if brand:
            queryset = (
                queryset.filter(
                    brand__slug=brand
                )
            )

        # -------------------------------------------------
        # Collection filters
        # -------------------------------------------------

        if query_bool(
            params.get(
                "new_arrival"
            )
        ):
            queryset = (
                queryset.filter(
                    is_new_arrival=True
                )
            )

        if query_bool(
            params.get(
                "clearance"
            )
        ):
            queryset = (
                queryset.filter(
                    is_clearance_sale=True
                )
            )

        if query_bool(
            params.get(
                "featured"
            )
        ):
            queryset = (
                queryset.filter(
                    is_featured=True
                )
            )

        if query_bool(
            params.get(
                "best_seller"
            )
        ):
            queryset = (
                queryset.filter(
                    is_best_seller=True
                )
            )

        if query_bool(
            params.get(
                "trending"
            )
        ):
            queryset = (
                queryset.filter(
                    is_trending=True
                )
            )

        if query_bool(
            params.get(
                "offer"
            )
        ):
            queryset = (
                queryset.filter(
                    is_offer=True
                )
            )

        # -------------------------------------------------
        # Price filters
        # -------------------------------------------------

        min_price = query_decimal(
            params.get(
                "min_price"
            )
        )

        max_price = query_decimal(
            params.get(
                "max_price"
            )
        )

        if min_price is not None:
            queryset = (
                queryset.filter(
                    price__gte=min_price
                )
            )

        if max_price is not None:
            queryset = (
                queryset.filter(
                    price__lte=max_price
                )
            )

        # -------------------------------------------------
        # Sorting
        # -------------------------------------------------

        ordering = (
            params
            .get(
                "ordering",
                "",
            )
            .strip()
        )

        allowed_ordering = {
            "price":
                "price",

            "-price":
                "-price",

            "price_asc":
                "price",

            "price_desc":
                "-price",

            "name":
                "name",

            "-name":
                "-name",

            "name_asc":
                "name",

            "name_desc":
                "-name",

            "created_at":
                "created_at",

            "-created_at":
                "-created_at",

            "newest":
                "-created_at",

            "oldest":
                "created_at",

            "rating":
                "approved_average_rating",

            "-rating":
                "-approved_average_rating",

            "rating_desc":
                "-approved_average_rating",
        }

        if ordering in allowed_ordering:
            queryset = (
                queryset.order_by(
                    allowed_ordering[
                        ordering
                    ],
                    "-created_at",
                )
            )

        return queryset.distinct()


# =========================================================
# Product Detail
# =========================================================

class ProductDetailView(
    generics.RetrieveAPIView
):
    serializer_class = (
        ProductDetailSerializer
    )

    permission_classes = (
        permissions.AllowAny,
    )

    lookup_field = "id"

    def get_queryset(self):
        return (
            base_product_queryset()
        )


# =========================================================
# Admin Product List / Create
# =========================================================

class AdminProductListCreateView(
    generics.ListCreateAPIView
):
    """
    GET:
        Return all products for admin,
        including inactive products.

    POST:
        Create product with optional nested variants.
    """

    serializer_class = (
        ProductAdminWriteSerializer
    )

    permission_classes = (
        IsAdminUserForProducts,
    )

    def get_queryset(self):
        queryset = (
            admin_product_queryset()
        )

        params = (
            self.request.query_params
        )

        search = (
            params
            .get(
                "search",
                "",
            )
            .strip()
        )

        if search:
            queryset = (
                queryset.filter(
                    Q(
                        name__icontains=search
                    )
                    | Q(
                        sku__icontains=search
                    )
                    | Q(
                        slug__icontains=search
                    )
                    | Q(
                        brand__name__icontains=search
                    )
                )
            )

        department = (
            params.get(
                "department"
            )
        )

        if department:
            queryset = (
                queryset.filter(
                    department__slug=department
                )
            )

        category = (
            params.get(
                "category"
            )
        )

        if category:
            queryset = (
                queryset.filter(
                    category__slug=category
                )
            )

        active = (
            params.get(
                "is_active"
            )
        )

        if active is not None:
            queryset = (
                queryset.filter(
                    is_active=query_bool(
                        active
                    )
                )
            )

        return queryset.distinct()


# =========================================================
# Admin Product Retrieve / Update / Delete
# =========================================================

class AdminProductDetailView(
    generics.RetrieveUpdateDestroyAPIView
):
    """
    GET:
        Retrieve a product for admin editing.

    PUT/PATCH:
        Update product and optionally nested variants.

    DELETE:
        Delete product and cascading variants/images.
    """

    serializer_class = (
        ProductAdminWriteSerializer
    )

    permission_classes = (
        IsAdminUserForProducts,
    )

    lookup_field = "id"

    def get_queryset(self):
        return (
            admin_product_queryset()
        )


# =========================================================
# Admin Variant List / Create
# =========================================================

class AdminProductVariantListCreateView(
    generics.ListCreateAPIView
):
    """
    GET:
        List variants.

    POST:
        Create a standalone product variant.

    Optional:
        ?product=1
    """

    serializer_class = (
        ProductVariantAdminSerializer
    )

    permission_classes = (
        IsAdminUserForProducts,
    )

    def get_queryset(self):
        queryset = (
            ProductVariant.objects
            .select_related(
                "product"
            )
            .order_by(
                "product",
                "color",
                "size",
            )
        )

        product_id = (
            self.request
            .query_params
            .get(
                "product"
            )
        )

        if product_id:
            queryset = (
                queryset.filter(
                    product_id=product_id
                )
            )

        return queryset


# =========================================================
# Admin Variant Retrieve / Update / Delete
# =========================================================

class AdminProductVariantDetailView(
    generics.RetrieveUpdateDestroyAPIView
):
    serializer_class = (
        ProductVariantAdminSerializer
    )

    permission_classes = (
        IsAdminUserForProducts,
    )

    lookup_field = "id"

    queryset = (
        ProductVariant.objects
        .select_related(
            "product"
        )
        .all()
    )


# =========================================================
# Admin Image List / Create
# =========================================================

class AdminProductImageListCreateView(
    generics.ListCreateAPIView
):
    """
    GET:
        List gallery images.

    POST:
        Upload/create a product gallery image.

    Optional:
        ?product=1
    """

    serializer_class = (
        ProductImageAdminSerializer
    )

    permission_classes = (
        IsAdminUserForProducts,
    )

    def get_queryset(self):
        queryset = (
            ProductImage.objects
            .select_related(
                "product"
            )
            .order_by(
                "product",
                "order",
                "id",
            )
        )

        product_id = (
            self.request
            .query_params
            .get(
                "product"
            )
        )

        if product_id:
            queryset = (
                queryset.filter(
                    product_id=product_id
                )
            )

        return queryset


# =========================================================
# Admin Image Retrieve / Update / Delete
# =========================================================

class AdminProductImageDetailView(
    generics.RetrieveUpdateDestroyAPIView
):
    serializer_class = (
        ProductImageAdminSerializer
    )

    permission_classes = (
        IsAdminUserForProducts,
    )

    lookup_field = "id"

    queryset = (
        ProductImage.objects
        .select_related(
            "product"
        )
        .all()
    )


# =========================================================
# Admin Bulk Product Upload
# =========================================================

class AdminBulkProductUploadView(
    APIView
):
    """
    Upload an Excel file from the React admin panel.

    Accepted multipart field names:
        file
        excel_file

    Required Excel columns:
        sku
        name
        brand
        department
        category
        price

    Optional columns:
        slug
        subcategory
        description
        old_price
        is_active
        is_featured
        is_best_seller
        is_trending
        is_new_arrival
        is_clearance_sale
        is_offer
        meta_title
        meta_description
    """

    permission_classes = (
        IsAdminUserForProducts,
    )

    parser_classes = (
        MultiPartParser,
        FormParser,
    )

    def post(
        self,
        request,
    ):
        excel_file = (
            request.FILES.get(
                "file"
            )
            or request.FILES.get(
                "excel_file"
            )
        )

        if not excel_file:
            return Response(
                {
                    "detail":
                        "Excel file is required."
                },
                status=(
                    status.HTTP_400_BAD_REQUEST
                ),
            )

        try:
            result = (
                self.import_products(
                    excel_file
                )
            )

        except Exception as error:
            return Response(
                {
                    "detail":
                        error_message(
                            error
                        )
                },
                status=(
                    status.HTTP_400_BAD_REQUEST
                ),
            )

        return Response(
            {
                "message":
                    "Bulk product upload completed.",
                **result,
            },
            status=status.HTTP_200_OK,
        )

    def import_products(
        self,
        excel_file,
    ):
        workbook = load_workbook(
            excel_file,
            read_only=True,
            data_only=True,
        )

        try:
            sheet = workbook.active

            rows = sheet.iter_rows(
                values_only=True
            )

            try:
                raw_headers = next(
                    rows
                )

            except StopIteration as error:
                raise ValueError(
                    "Excel file is empty."
                ) from error

            headers = [
                clean_text(
                    header
                ).lower()
                for header
                in raw_headers
            ]

            required_headers = {
                "sku",
                "name",
                "brand",
                "department",
                "category",
                "price",
            }

            missing_headers = (
                required_headers
                .difference(
                    headers
                )
            )

            if missing_headers:
                raise ValueError(
                    (
                        "Missing required columns: "
                        + ", ".join(
                            sorted(
                                missing_headers
                            )
                        )
                    )
                )

            created_count = 0
            updated_count = 0
            skipped_count = 0

            errors = []

            for (
                row_number,
                row,
            ) in enumerate(
                rows,
                start=2,
            ):
                row_data = dict(
                    zip(
                        headers,
                        row,
                    )
                )

                if not any(
                    value not in {
                        None,
                        "",
                    }
                    for value
                    in row_data.values()
                ):
                    continue

                try:
                    with transaction.atomic():
                        sku = clean_text(
                            row_data.get(
                                "sku"
                            )
                        )

                        name = clean_text(
                            row_data.get(
                                "name"
                            )
                        )

                        brand_name = clean_text(
                            row_data.get(
                                "brand"
                            )
                        )

                        department_name = (
                            clean_text(
                                row_data.get(
                                    "department"
                                )
                            )
                        )

                        category_name = (
                            clean_text(
                                row_data.get(
                                    "category"
                                )
                            )
                        )

                        subcategory_name = (
                            clean_text(
                                row_data.get(
                                    "subcategory"
                                )
                            )
                        )

                        if not all(
                            (
                                sku,
                                name,
                                brand_name,
                                department_name,
                                category_name,
                            )
                        ):
                            raise ValueError(
                                (
                                    "Required values "
                                    "are missing."
                                )
                            )

                        price = clean_decimal(
                            row_data.get(
                                "price"
                            )
                        )

                        if price is None:
                            raise ValueError(
                                "Invalid price."
                            )

                        if price < 0:
                            raise ValueError(
                                (
                                    "Price cannot "
                                    "be negative."
                                )
                            )

                        old_price = (
                            clean_decimal(
                                row_data.get(
                                    "old_price"
                                ),
                                allow_empty=True,
                            )
                        )

                        if (
                            old_price is not None
                            and old_price < price
                        ):
                            raise ValueError(
                                (
                                    "Old price cannot "
                                    "be lower than "
                                    "selling price."
                                )
                            )

                        # ---------------------------------
                        # Brand
                        # ---------------------------------

                        brand = (
                            Brand.objects
                            .filter(
                                name__iexact=(
                                    brand_name
                                )
                            )
                            .first()
                        )

                        if brand is None:
                            brand = Brand(
                                name=brand_name,
                                slug=(
                                    unique_slug_for_model(
                                        Brand,
                                        brand_name,
                                    )
                                ),
                                is_active=True,
                            )

                            brand.full_clean()
                            brand.save()

                        # ---------------------------------
                        # Department
                        # ---------------------------------

                        department = (
                            Department.objects
                            .filter(
                                name__iexact=(
                                    department_name
                                )
                            )
                            .first()
                        )

                        if department is None:
                            department = (
                                Department(
                                    name=(
                                        department_name
                                    ),
                                    slug=(
                                        unique_slug_for_model(
                                            Department,
                                            department_name,
                                        )
                                    ),
                                    is_active=True,
                                    show_in_navbar=True,
                                )
                            )

                            department.full_clean()
                            department.save()

                        # ---------------------------------
                        # Category
                        # ---------------------------------

                        category = (
                            Category.objects
                            .filter(
                                department=(
                                    department
                                ),
                                name__iexact=(
                                    category_name
                                ),
                            )
                            .first()
                        )

                        if category is None:
                            category = Category(
                                department=(
                                    department
                                ),
                                name=(
                                    category_name
                                ),
                                slug=(
                                    unique_slug_for_model(
                                        Category,
                                        (
                                            f"{department_name}-"
                                            f"{category_name}"
                                        ),
                                    )
                                ),
                                is_active=True,
                                show_in_navbar=True,
                            )

                            category.full_clean()
                            category.save()

                        # ---------------------------------
                        # Subcategory
                        # ---------------------------------

                        subcategory = None

                        if subcategory_name:
                            subcategory = (
                                SubCategory.objects
                                .filter(
                                    category=(
                                        category
                                    ),
                                    name__iexact=(
                                        subcategory_name
                                    ),
                                )
                                .first()
                            )

                            if subcategory is None:
                                subcategory = (
                                    SubCategory(
                                        category=(
                                            category
                                        ),
                                        name=(
                                            subcategory_name
                                        ),
                                        slug=(
                                            unique_slug_for_model(
                                                SubCategory,
                                                (
                                                    f"{department_name}-"
                                                    f"{category_name}-"
                                                    f"{subcategory_name}"
                                                ),
                                            )
                                        ),
                                        is_active=True,
                                        show_in_navbar=True,
                                    )
                                )

                                subcategory.full_clean()
                                subcategory.save()

                        # ---------------------------------
                        # Product
                        # ---------------------------------

                        product = (
                            Product.objects
                            .filter(
                                sku__iexact=sku
                            )
                            .first()
                        )

                        created = (
                            product is None
                        )

                        if created:
                            product = Product(
                                sku=sku
                            )

                        supplied_slug = (
                            clean_text(
                                row_data.get(
                                    "slug"
                                )
                            )
                        )

                        if supplied_slug:
                            supplied_slug = (
                                slugify(
                                    supplied_slug
                                )
                            )

                            slug_queryset = (
                                Product.objects
                                .filter(
                                    slug=(
                                        supplied_slug
                                    )
                                )
                            )

                            if product.pk:
                                slug_queryset = (
                                    slug_queryset
                                    .exclude(
                                        pk=product.pk
                                    )
                                )

                            if slug_queryset.exists():
                                raise ValueError(
                                    (
                                        "Product slug "
                                        "already exists: "
                                        f"{supplied_slug}"
                                    )
                                )

                            product.slug = (
                                supplied_slug
                            )

                        elif (
                            created
                            or not product.slug
                        ):
                            product.slug = (
                                unique_slug_for_model(
                                    Product,
                                    (
                                        f"{name}-{sku}"
                                    ),
                                    exclude_id=(
                                        product.pk
                                        if product.pk
                                        else None
                                    ),
                                )
                            )

                        product.sku = sku
                        product.name = name
                        product.brand = brand

                        product.department = (
                            department
                        )

                        product.category = (
                            category
                        )

                        product.subcategory = (
                            subcategory
                        )

                        product.description = (
                            clean_text(
                                row_data.get(
                                    "description"
                                )
                            )
                        )

                        product.price = price

                        product.old_price = (
                            old_price
                        )

                        product.is_active = (
                            clean_boolean(
                                row_data.get(
                                    "is_active"
                                ),
                                default=True,
                            )
                        )

                        product.is_featured = (
                            clean_boolean(
                                row_data.get(
                                    "is_featured"
                                )
                            )
                        )

                        product.is_best_seller = (
                            clean_boolean(
                                row_data.get(
                                    "is_best_seller"
                                )
                            )
                        )

                        product.is_trending = (
                            clean_boolean(
                                row_data.get(
                                    "is_trending"
                                )
                            )
                        )

                        product.is_new_arrival = (
                            clean_boolean(
                                row_data.get(
                                    "is_new_arrival"
                                )
                            )
                        )

                        product.is_clearance_sale = (
                            clean_boolean(
                                row_data.get(
                                    "is_clearance_sale"
                                )
                            )
                        )

                        product.is_offer = (
                            clean_boolean(
                                row_data.get(
                                    "is_offer"
                                )
                            )
                        )

                        product.meta_title = (
                            clean_text(
                                row_data.get(
                                    "meta_title"
                                )
                            )
                        )

                        product.meta_description = (
                            clean_text(
                                row_data.get(
                                    "meta_description"
                                )
                            )
                        )

                        product.full_clean()
                        product.save()

                        if created:
                            created_count += 1

                        else:
                            updated_count += 1

                except Exception as error:
                    skipped_count += 1

                    errors.append(
                        {
                            "row":
                                row_number,
                            "sku":
                                clean_text(
                                    row_data.get(
                                        "sku"
                                    )
                                ),
                            "error":
                                error_message(
                                    error
                                ),
                        }
                    )

            return {
                "created":
                    created_count,
                "updated":
                    updated_count,
                "skipped":
                    skipped_count,
                "errors":
                    errors[:100],
            }

        finally:
            workbook.close()


# =========================================================
# Admin Bulk Product Variant Upload
# =========================================================

class AdminBulkVariantUploadView(
    APIView
):
    """
    Upload product variants through Excel.

    Accepted multipart field names:
        file
        excel_file

    Required columns:
        product_sku
        variant_sku
        color
        size
        stock

    Optional:
        color_code
        is_active
    """

    permission_classes = (
        IsAdminUserForProducts,
    )

    parser_classes = (
        MultiPartParser,
        FormParser,
    )

    def post(
        self,
        request,
    ):
        excel_file = (
            request.FILES.get(
                "file"
            )
            or request.FILES.get(
                "excel_file"
            )
        )

        if not excel_file:
            return Response(
                {
                    "detail":
                        "Excel file is required."
                },
                status=(
                    status.HTTP_400_BAD_REQUEST
                ),
            )

        try:
            result = (
                self.import_variants(
                    excel_file
                )
            )

        except Exception as error:
            return Response(
                {
                    "detail":
                        error_message(
                            error
                        )
                },
                status=(
                    status.HTTP_400_BAD_REQUEST
                ),
            )

        return Response(
            {
                "message":
                    "Bulk variant upload completed.",
                **result,
            },
            status=status.HTTP_200_OK,
        )

    def import_variants(
        self,
        excel_file,
    ):
        workbook = load_workbook(
            excel_file,
            read_only=True,
            data_only=True,
        )

        try:
            sheet = workbook.active

            rows = sheet.iter_rows(
                values_only=True
            )

            try:
                raw_headers = next(
                    rows
                )

            except StopIteration as error:
                raise ValueError(
                    "Excel file is empty."
                ) from error

            headers = [
                clean_text(
                    header
                ).lower()
                for header
                in raw_headers
            ]

            required_headers = {
                "product_sku",
                "variant_sku",
                "color",
                "size",
                "stock",
            }

            missing_headers = (
                required_headers
                .difference(
                    headers
                )
            )

            if missing_headers:
                raise ValueError(
                    (
                        "Missing required columns: "
                        + ", ".join(
                            sorted(
                                missing_headers
                            )
                        )
                    )
                )

            created_count = 0
            updated_count = 0
            skipped_count = 0

            errors = []

            for (
                row_number,
                row,
            ) in enumerate(
                rows,
                start=2,
            ):
                row_data = dict(
                    zip(
                        headers,
                        row,
                    )
                )

                if not any(
                    value not in {
                        None,
                        "",
                    }
                    for value
                    in row_data.values()
                ):
                    continue

                try:
                    with transaction.atomic():
                        product_sku = (
                            clean_text(
                                row_data.get(
                                    "product_sku"
                                )
                            )
                        )

                        variant_sku = (
                            clean_text(
                                row_data.get(
                                    "variant_sku"
                                )
                            )
                        )

                        color = clean_text(
                            row_data.get(
                                "color"
                            )
                        )

                        color_code = (
                            clean_text(
                                row_data.get(
                                    "color_code"
                                )
                            )
                        )

                        size = clean_text(
                            row_data.get(
                                "size"
                            )
                        )

                        if not all(
                            (
                                product_sku,
                                variant_sku,
                                color,
                                size,
                            )
                        ):
                            raise ValueError(
                                (
                                    "Required values "
                                    "are missing."
                                )
                            )

                        product = (
                            Product.objects
                            .filter(
                                sku__iexact=(
                                    product_sku
                                )
                            )
                            .first()
                        )

                        if product is None:
                            raise ValueError(
                                (
                                    "Product SKU "
                                    f"{product_sku} "
                                    "not found."
                                )
                            )

                        stock_raw = (
                            row_data.get(
                                "stock"
                            )
                        )

                        try:
                            stock = int(
                                stock_raw or 0
                            )

                        except (
                            TypeError,
                            ValueError,
                        ) as error:
                            raise ValueError(
                                (
                                    "Invalid stock "
                                    "value."
                                )
                            ) from error

                        if stock < 0:
                            raise ValueError(
                                (
                                    "Stock cannot "
                                    "be negative."
                                )
                            )

                        is_active = (
                            clean_boolean(
                                row_data.get(
                                    "is_active"
                                ),
                                default=True,
                            )
                        )

                        # ---------------------------------
                        # Prefer variant SKU
                        # ---------------------------------

                        variant = (
                            ProductVariant.objects
                            .filter(
                                sku__iexact=(
                                    variant_sku
                                )
                            )
                            .first()
                        )

                        # ---------------------------------
                        # Otherwise same product/color/size
                        # ---------------------------------

                        if variant is None:
                            variant = (
                                ProductVariant.objects
                                .filter(
                                    product=(
                                        product
                                    ),
                                    color__iexact=(
                                        color
                                    ),
                                    size__iexact=(
                                        size
                                    ),
                                )
                                .first()
                            )

                        created = (
                            variant is None
                        )

                        if created:
                            variant = (
                                ProductVariant()
                            )

                        variant.product = (
                            product
                        )

                        variant.sku = (
                            variant_sku
                        )

                        variant.color = (
                            color
                        )

                        variant.color_code = (
                            color_code
                        )

                        variant.size = (
                            size
                        )

                        variant.stock = (
                            stock
                        )

                        variant.is_active = (
                            is_active
                        )

                        variant.full_clean()
                        variant.save()

                        if created:
                            created_count += 1

                        else:
                            updated_count += 1

                except Exception as error:
                    skipped_count += 1

                    errors.append(
                        {
                            "row":
                                row_number,
                            "product_sku":
                                clean_text(
                                    row_data.get(
                                        "product_sku"
                                    )
                                ),
                            "variant_sku":
                                clean_text(
                                    row_data.get(
                                        "variant_sku"
                                    )
                                ),
                            "error":
                                error_message(
                                    error
                                ),
                        }
                    )

            return {
                "created":
                    created_count,
                "updated":
                    updated_count,
                "skipped":
                    skipped_count,
                "errors":
                    errors[:100],
            }

        finally:
            workbook.close()


# =========================================================
# Admin Bulk Product Image ZIP Upload
# =========================================================

class AdminBulkProductImageUploadView(
    APIView
):
    """
    Upload product images using one ZIP file.

    Accepted multipart fields:
        file
        zip_file

    Naming convention:

        SKU.jpg
            -> Product main image

        SKU_1.jpg
            -> Gallery image order 1

        SKU_2.jpg
            -> Gallery image order 2

    Supported:
        jpg
        jpeg
        png
        webp
    """

    permission_classes = (
        IsAdminUserForProducts,
    )

    parser_classes = (
        MultiPartParser,
        FormParser,
    )

    def post(
        self,
        request,
    ):
        zip_file = (
            request.FILES.get(
                "file"
            )
            or request.FILES.get(
                "zip_file"
            )
        )

        if not zip_file:
            return Response(
                {
                    "detail":
                        "ZIP file is required."
                },
                status=(
                    status.HTTP_400_BAD_REQUEST
                ),
            )

        try:
            result = (
                self.import_images(
                    zip_file
                )
            )

        except Exception as error:
            return Response(
                {
                    "detail":
                        error_message(
                            error
                        )
                },
                status=(
                    status.HTTP_400_BAD_REQUEST
                ),
            )

        return Response(
            {
                "message":
                    "Bulk image upload completed.",
                **result,
            },
            status=status.HTTP_200_OK,
        )

    @transaction.atomic
    def import_images(
        self,
        zip_file,
    ):
        allowed_extensions = {
            ".jpg",
            ".jpeg",
            ".png",
            ".webp",
        }

        main_image_count = 0
        gallery_image_count = 0
        skipped_count = 0

        errors = []

        try:
            archive = ZipFile(
                zip_file
            )

        except BadZipFile as error:
            raise ValueError(
                (
                    "Uploaded file is "
                    "not a valid ZIP file."
                )
            ) from error

        products_by_sku = {
            product.sku
            .strip()
            .lower():
                product
            for product
            in Product.objects.all()
            if product.sku
        }

        try:
            for zip_info in (
                archive.infolist()
            ):
                if zip_info.is_dir():
                    continue

                original_name = (
                    Path(
                        zip_info.filename
                    ).name
                )

                if not original_name:
                    continue

                if original_name.startswith(
                    "."
                ):
                    continue

                file_path = Path(
                    original_name
                )

                extension = (
                    file_path
                    .suffix
                    .lower()
                )

                if (
                    extension
                    not in allowed_extensions
                ):
                    skipped_count += 1

                    errors.append(
                        {
                            "file":
                                original_name,
                            "error":
                                (
                                    "Unsupported image "
                                    "file type."
                                ),
                        }
                    )

                    continue

                file_stem = (
                    file_path.stem
                )

                product = None
                gallery_order = None

                exact_sku = (
                    file_stem.lower()
                )

                if (
                    exact_sku
                    in products_by_sku
                ):
                    product = (
                        products_by_sku[
                            exact_sku
                        ]
                    )

                elif "_" in file_stem:
                    (
                        possible_sku,
                        possible_order,
                    ) = file_stem.rsplit(
                        "_",
                        1,
                    )

                    if (
                        possible_order
                        .isdigit()
                        and (
                            possible_sku
                            .lower()
                            in products_by_sku
                        )
                    ):
                        product = (
                            products_by_sku[
                                possible_sku
                                .lower()
                            ]
                        )

                        gallery_order = (
                            int(
                                possible_order
                            )
                        )

                if product is None:
                    skipped_count += 1

                    errors.append(
                        {
                            "file":
                                original_name,
                            "error":
                                (
                                    "Matching product "
                                    "SKU not found."
                                ),
                        }
                    )

                    continue

                try:
                    image_bytes = (
                        archive.read(
                            zip_info
                        )
                    )

                    if not image_bytes:
                        raise ValueError(
                            "Image file is empty."
                        )

                    image_content = (
                        ContentFile(
                            image_bytes
                        )
                    )

                    # ---------------------------------
                    # Main Image
                    # ---------------------------------

                    if gallery_order is None:
                        if product.main_image:
                            product.main_image.delete(
                                save=False
                            )

                        product.main_image.save(
                            original_name,
                            image_content,
                            save=True,
                        )

                        main_image_count += 1

                    # ---------------------------------
                    # Gallery Image
                    # ---------------------------------

                    else:
                        existing_image = (
                            ProductImage.objects
                            .filter(
                                product=(
                                    product
                                ),
                                order=(
                                    gallery_order
                                ),
                            )
                            .first()
                        )

                        if existing_image:
                            if (
                                existing_image
                                .image
                            ):
                                (
                                    existing_image
                                    .image
                                    .delete(
                                        save=False
                                    )
                                )

                            existing_image.alt_text = (
                                product.name
                            )

                            existing_image.order = (
                                gallery_order
                            )

                            (
                                existing_image
                                .image
                                .save(
                                    original_name,
                                    image_content,
                                    save=False,
                                )
                            )

                            existing_image.full_clean()
                            existing_image.save()

                        else:
                            gallery_image = (
                                ProductImage(
                                    product=(
                                        product
                                    ),
                                    alt_text=(
                                        product.name
                                    ),
                                    order=(
                                        gallery_order
                                    ),
                                )
                            )

                            (
                                gallery_image
                                .image
                                .save(
                                    original_name,
                                    image_content,
                                    save=False,
                                )
                            )

                            gallery_image.full_clean()
                            gallery_image.save()

                        gallery_image_count += 1

                except Exception as error:
                    skipped_count += 1

                    errors.append(
                        {
                            "file":
                                original_name,
                            "error":
                                error_message(
                                    error
                                ),
                        }
                    )

        finally:
            archive.close()

        return {
            "main_images":
                main_image_count,
            "gallery_images":
                gallery_image_count,
            "skipped":
                skipped_count,
            "errors":
                errors[:100],
        }